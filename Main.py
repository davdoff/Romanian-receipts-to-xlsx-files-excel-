"""
Romanian Fiscal Receipt Processor — Claude Vision API → Excel
-------------------------------------------------------------
Usage:
  1. pip install anthropic openpyxl pillow
  2. Set ANTHROPIC_API_KEY in your environment
  3. Put receipt images in a folder (jpg, png, webp, gif)
  4. Run: python receipt_processor.py --folder ./receipts --output receipts.xlsx

Efficiency notes:
  - Assistant prefill forces JSON output with zero preamble (saves output tokens)
  - System prompt is terse and schema-first (reduces input tokens vs verbose instructions)
  - max_tokens set to 1500 — enough for a full receipt with many items
  - Use --skip-errors to continue past bad/blurry images without stopping the batch
"""

import anthropic
import base64
import json
import os
import re
import argparse
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from dotenv import load_dotenv

load_dotenv()

SUPPORTED_EXTENSIONS = {".jpg", ".jpeg", ".png", ".webp", ".gif"}

# ── Prompt design notes ───────────────────────────────────────────────────────
# 1. Schema-first: the model sees the output shape before any rules.
#    This anchors its attention on structure rather than prose instructions.
# 2. Rules are short and imperative — no padding, no examples in the system prompt.
#    Examples cost input tokens on every single call; save them for ambiguous edge
#    cases you handle in post-processing instead.
# 3. Assistant prefill (the {"role":"assistant","content":"{"} message) is the
#    biggest efficiency win: the model skips any preamble and outputs pure JSON
#    from token 1. No markdown fences to strip, no "Here is the extracted data:"
# 4. VAT letters on receipts are unreliable (some vendors print B at 21%).
#    We always extract the numeric rate and remap to A=21%, B=11%, C=5% in
#    normalize_vat_by_rate(). If a rate doesn't match any known bracket, it's dropped.
# ─────────────────────────────────────────────────────────────────────────────

# Maps the actual VAT rate (%) to the canonical output category letter.
RATE_TO_CATEGORY = {21: "A", 11: "B", 5: "C"}

RECEIPT_SYSTEM_PROMPT = """You extract data from Romanian fiscal receipts (bonuri fiscale).
Output ONLY a valid JSON object matching this exact schema — no markdown, no text outside the JSON.

SCHEMA:
{
  "receipt_number": "string or null",
  "date": "YYYY-MM-DD or null",
  "payment_method": "Cash | Card | null",
  "vendor": {
    "name": "string or null",
    "address": "string or null",
    "cui": "string or null"
  },
  "client": {
    "name": "string or null",
    "cif": "string or null"
  },
  "vat_breakdown": [
    {"rate": number, "vat_amount": number, "total": number}
  ],
  "total_with_vat": number or null,
  "total_vat": number or null,
  "notes": "string or null"
}

RULES:
- All amounts as plain numbers (no currency symbols), currency is always RON
- vat_breakdown: one entry per VAT rate that appears on the receipt; read the rate % directly from the receipt — never assume a rate; on Romanian receipts the TVA table columns are "TVA | VALOARE | TOTAL" — VALOARE is the vat_amount and TOTAL is the category total including VAT; ignore the letter (A/B/C) on the receipt
- client: use null for client name if it is not included but check if the cif is specified, if so add it; Romanian CIF format is the letters "RO" followed by 2–10 digits (e.g. RO12345678) — read every digit carefully, do not skip any
- vendor.cui: the seller's fiscal identification code — may appear as "CUI", "CIF", or "Cod identificare fiscala" on the receipt; same RO + 2–10 digit format applies
- receipt_number: found at the bottom of the receipt, may appear as "BF", "Nr. bon", "Numar bon", or "Numar bon fiscal"
- If a field is not visible or not applicable, use null
- Ignore the items, the important infromation is before the items start with and after the total"""

FACTURA_SYSTEM_PROMPT = """You extract data from Romanian fiscal invoices (facturi).
Output ONLY a valid JSON object matching this exact schema — no markdown, no text outside the JSON.

SCHEMA:
{
  "invoice_number": "string or null",
  "date": "YYYY-MM-DD or null",
  "vendor": {"name": "string or null"},
  "client": {"name": "string or null"},
  "contract_number": "string or null",
  "livrare": "string or null",
  "total_plata": number or null,
  "relevant": true or false
}

RULES:
- invoice_number: the invoice series + number, e.g. "26M02607777"; may appear as "Factura nr.", "Nr. factura", or in a header box
- date: the issue date (data emiterii), format YYYY-MM-DD
- contract_number: may appear as "Nr. contract", "Numar contract", "Contract nr."
- livrare: the delivery or service period string as it appears on the invoice (e.g. "01.08.2024 - 31.08.2024"); may also appear as "Perioada de facturare" or "Interval"
- total_plata: the final total amount due (Total de plata / Total plata), as a plain number — no currency symbols, currency is always RON
- relevant: set to true only if the vendor name contains "PPC Energie"; set to false for any other vendor
- If a field is not visible or not applicable, use null"""



_CIF_RE = re.compile(r'^RO\d{2,10}$', re.IGNORECASE)


def validate_receipt(data: dict) -> list[str]:
    """Check key fields for obvious format errors and return a list of warning strings."""
    warnings = []
    cif = (data.get("client") or {}).get("cif")
    if cif and not _CIF_RE.match(str(cif).strip()):
        warnings.append(f"CIF CLIENT format suspect: '{cif}' (expected RO + 2-10 digits)")
    cui = (data.get("vendor") or {}).get("cui")
    if cui and not _CIF_RE.match(str(cui).strip()):
        warnings.append(f"CUI vendor format suspect: '{cui}' (expected RO + 2-10 digits)")
    return warnings


def normalize_vat_by_rate(vat_breakdown) -> dict:
    """Index VAT entries by canonical letter (A/B/C) keyed on rate, compute base = total - vat_amount."""
    normalized = {}
    entries = vat_breakdown if isinstance(vat_breakdown, list) else list((vat_breakdown or {}).values())
    for entry in entries:
        if not entry:
            continue
        rate = entry.get("rate")
        if rate is None:
            continue
        cat = RATE_TO_CATEGORY.get(int(round(float(rate))))
        if cat:
            vat_amount = entry.get("vat_amount")
            total = entry.get("total")
            if vat_amount is not None and total is not None:
                entry["base"] = round(total - vat_amount, 2)
            normalized[cat] = entry
    return normalized


def clean_bleedthrough(img_bgr):
    """Remove back-side logo bleed-through via adaptive threshold + morphological opening.

    blockSize=31, C=15 — tune blockSize toward 51 or C toward 20 if logos still show;
    lower C toward 10 if real text starts getting cut off.
    """
    import cv2
    import numpy as np
    gray = cv2.cvtColor(img_bgr, cv2.COLOR_BGR2GRAY)
    binary = cv2.adaptiveThreshold(
        gray, 255,
        cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
        cv2.THRESH_BINARY,
        blockSize=31, C=15,
    )
    kernel = np.ones((2, 2), np.uint8)
    return cv2.morphologyEx(binary, cv2.MORPH_OPEN, kernel)


def ocr_with_vision(image_path: Path) -> str:
    from google.cloud import vision
    api_key = os.environ.get("GOOGLE_VISION_API_KEY")
    client_options = {"api_key": api_key} if api_key else {}
    vision_client = vision.ImageAnnotatorClient(client_options=client_options)
    with open(image_path, "rb") as f:
        content = f.read()
    image = vision.Image(content=content)
    response = vision_client.document_text_detection(image=image)
    if response.error.message:
        raise RuntimeError(f"Google Vision error: {response.error.message}")
    return response.full_text_annotation.text


def ocr_pil_with_vision(pil_img) -> str:
    import io
    from google.cloud import vision
    api_key = os.environ.get("GOOGLE_VISION_API_KEY")
    client_options = {"api_key": api_key} if api_key else {}
    vision_client = vision.ImageAnnotatorClient(client_options=client_options)
    buf = io.BytesIO()
    pil_img.save(buf, format="JPEG")
    image = vision.Image(content=buf.getvalue())
    response = vision_client.document_text_detection(image=image)
    if response.error.message:
        raise RuntimeError(f"Google Vision error: {response.error.message}")
    return response.full_text_annotation.text


def encode_image(image_path: Path) -> tuple[str, str]:
    import cv2
    img = cv2.imread(str(image_path))
    if img is not None:
        cleaned = clean_bleedthrough(img)
        _, buf = cv2.imencode(".jpg", cleaned, [cv2.IMWRITE_JPEG_QUALITY, 95])
        return base64.standard_b64encode(buf.tobytes()).decode("utf-8"), "image/jpeg"
    # fallback for formats cv2 can't read (e.g. gif)
    ext = image_path.suffix.lower()
    media_type_map = {".jpg": "image/jpeg", ".jpeg": "image/jpeg",
                      ".png": "image/png", ".webp": "image/webp", ".gif": "image/gif"}
    with open(image_path, "rb") as f:
        return base64.standard_b64encode(f.read()).decode("utf-8"), media_type_map.get(ext, "image/jpeg")


def encode_pil_image(pil_img) -> tuple[str, str]:
    import cv2
    import numpy as np
    img_bgr = cv2.cvtColor(np.array(pil_img.convert("RGB")), cv2.COLOR_RGB2BGR)
    cleaned = clean_bleedthrough(img_bgr)
    _, buf = cv2.imencode(".jpg", cleaned, [cv2.IMWRITE_JPEG_QUALITY, 95])
    return base64.standard_b64encode(buf.tobytes()).decode("utf-8"), "image/jpeg"


def _call_api(client: anthropic.Anthropic, ocr_text: str, system_prompt: str) -> dict:
    response = client.messages.create(
        model="claude-haiku-4-5-20251001",
        max_tokens=4096,
        system=system_prompt,
        messages=[
            {
                "role": "user",
                "content": f"Extract all data from this Romanian fiscal receipt:\n\n{ocr_text}",
            },
        ],
    )
    raw = response.content[0].text.strip()
    if raw.startswith("```"):
        raw = raw.split("```")[1]
        if raw.startswith("json"):
            raw = raw[4:]
    return json.loads(raw.strip())


def extract_receipt(client: anthropic.Anthropic, image_path: Path) -> dict:
    ocr_text = ocr_with_vision(image_path)
    return _call_api(client, ocr_text, RECEIPT_SYSTEM_PROMPT)


def extract_receipt_from_pil(client: anthropic.Anthropic, pil_img) -> dict:
    ocr_text = ocr_pil_with_vision(pil_img)
    return _call_api(client, ocr_text, RECEIPT_SYSTEM_PROMPT)


def extract_factura(client: anthropic.Anthropic, image_path: Path) -> dict:
    ocr_text = ocr_with_vision(image_path)
    return _call_api(client, ocr_text, FACTURA_SYSTEM_PROMPT)


def extract_factura_from_pil(client: anthropic.Anthropic, pil_img) -> dict:
    ocr_text = ocr_pil_with_vision(pil_img)
    return _call_api(client, ocr_text, FACTURA_SYSTEM_PROMPT)


def style_header(cell, header_fill, header_font, center, border):
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = center
    cell.border = border


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w


def build_excel(records: list[dict], output_path: str):
    wb = openpyxl.Workbook()

    header_fill = PatternFill("solid", start_color="2F5496")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell_font   = Font(name="Arial", size=10)
    bold_font   = Font(bold=True, name="Arial", size=10)
    alt_fill    = PatternFill("solid", start_color="DCE6F1")
    border      = Border(
        bottom=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
    )
    center = Alignment(horizontal="center", vertical="center")
    money  = '#,##0.00'

    # ── Sheet 1: Summary ──────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Summary"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20

    summary_headers = [
        "File",
        "Nr. Bon Fiscal",
        "Data",
        "Furnizor",
        "CUI Furnizor",
        "Adresa Furnizor",
        "Client",
        "CIF Client",
        "Cash / Card",
        "Baza 21%",
        "TVA 21%",
        "Baza 11%",
        "TVA 11%",
        "Baza 5%",
        "TVA 5%",
        "Total cu TVA",
        "Total TVA",
        "Parse Error",
    ]

    for col, h in enumerate(summary_headers, 1):
        style_header(ws.cell(row=1, column=col, value=h), header_fill, header_font, center, border)

    set_col_widths(ws, [22, 16, 12, 28, 14, 36, 28, 14, 10, 14, 12, 14, 12, 14, 12, 18, 16, 24])
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(summary_headers))}1"

    # money columns: Baza A=10, TVA A=11, Baza B=12, TVA B=13, Baza C=14, TVA C=15, Total=16, TotalTVA=17
    money_cols = {10, 11, 12, 13, 14, 15, 16, 17}

    for row_idx, rec in enumerate(records, 2):
        d   = rec.get("data", {})
        v   = d.get("vendor", {}) or {}
        cl  = d.get("client", {}) or {}
        vat = d.get("vat_breakdown", {}) or {}
        err = rec.get("error", "")
        fill = alt_fill if row_idx % 2 == 0 else None

        raw_date = d.get("date")
        if raw_date:
            try:
                from datetime import datetime
                raw_date = datetime.strptime(raw_date, "%Y-%m-%d").strftime("%d/%m/%Y")
            except Exception:
                pass

        row_values = [
            rec.get("filename"),
            d.get("receipt_number"),
            raw_date,
            v.get("name"),
            v.get("cui"),
            v.get("address"),
            cl.get("name"),
            cl.get("cif"),
            d.get("payment_method"),
        ]
        for cat in ("A", "B", "C"):
            entry = vat.get(cat) or {}
            row_values.append(entry.get("base"))
            row_values.append(entry.get("vat_amount"))
        row_values.append(d.get("total_with_vat"))
        row_values.append(d.get("total_vat"))
        row_values.append(err)

        for col, val in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = cell_font
            cell.border = border
            if fill:
                cell.fill = fill
            if col in money_cols and isinstance(val, (int, float)):
                cell.number_format = money

    # Totals row
    total_row = len(records) + 2
    ws.cell(row=total_row, column=9, value="TOTAL").font = bold_font
    col_letters = {10: "J", 11: "K", 12: "L", 13: "M", 14: "N", 15: "O", 16: "P", 17: "Q"}
    for col, col_letter in col_letters.items():
        cell = ws.cell(row=total_row, column=col,
                       value=f"=SUM({col_letter}2:{col_letter}{total_row-1})")
        cell.font = bold_font
        cell.number_format = money

    # ── Sheet 3: Line Items (disabled) ───────────────────────────────────────
    # ws3 = wb.create_sheet("Produse")
    # ws3.freeze_panes = "A2"
    # ws3.row_dimensions[1].height = 20
    # item_headers = [
    #     "Nr. Bon Fiscal",
    #     "Produs", "Cantitate", "Pret Unitar", "Total Linie", "Categorie TVA",
    # ]
    # for col, h in enumerate(item_headers, 1):
    #     style_header(ws3.cell(row=1, column=col, value=h), header_fill, header_font, center, border)
    # set_col_widths(ws3, [16, 36, 10, 16, 16, 14])
    # ws3.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(item_headers))}1"
    # item_row = 2
    # for rec in records:
    #     d = rec.get("data", {})
    #     for item in d.get("items", []):
    #         fill = alt_fill if item_row % 2 == 0 else None
    #         row_values = [
    #             d.get("receipt_number"),
    #             item.get("name"),
    #             item.get("qty"),
    #             item.get("unit_price"),
    #             item.get("line_total"),
    #             item.get("vat_category"),
    #         ]
    #         for col, val in enumerate(row_values, 1):
    #             cell = ws3.cell(row=item_row, column=col, value=val)
    #             cell.font = cell_font
    #             cell.border = border
    #             if fill:
    #                 cell.fill = fill
    #             if col in (4, 5) and isinstance(val, (int, float)):
    #                 cell.number_format = money
    #         item_row += 1

    wb.save(output_path)
    print(f"\n✅ Saved → {output_path}")
    print(f"   {len(records)} receipts")


def build_excel_facturi(records: list[dict], output_path: str):
    wb = openpyxl.Workbook()

    header_fill = PatternFill("solid", start_color="2F5496")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    cell_font   = Font(name="Arial", size=10)
    bold_font   = Font(bold=True, name="Arial", size=10)
    alt_fill    = PatternFill("solid", start_color="DCE6F1")
    border      = Border(
        bottom=Side(style="thin", color="BFBFBF"),
        right=Side(style="thin", color="BFBFBF"),
    )
    center = Alignment(horizontal="center", vertical="center")
    money  = '#,##0.00'

    ws = wb.active
    ws.title = "Facturi"
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 20

    headers = [
        "File",
        "Factura Nr",
        "Data Emiterii",
        "Furnizor",
        "Client",
        "Numar Contract",
        "Livrare",
        "Total Plata",
        "Parse Error",
    ]

    for col, h in enumerate(headers, 1):
        style_header(ws.cell(row=1, column=col, value=h), header_fill, header_font, center, border)

    set_col_widths(ws, [22, 18, 14, 28, 28, 18, 28, 16, 24])
    ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(len(headers))}1"

    money_col = 8  # Total Plata

    for row_idx, rec in enumerate(records, 2):
        d  = rec.get("data", {})
        v  = d.get("vendor", {}) or {}
        cl = d.get("client", {}) or {}
        err = rec.get("error", "")
        fill = alt_fill if row_idx % 2 == 0 else None

        raw_date = d.get("date")
        if raw_date:
            try:
                from datetime import datetime
                raw_date = datetime.strptime(raw_date, "%Y-%m-%d").strftime("%d/%m/%Y")
            except Exception:
                pass

        row_values = [
            rec.get("filename"),
            d.get("invoice_number"),
            raw_date,
            v.get("name"),
            cl.get("name"),
            d.get("contract_number"),
            d.get("livrare"),
            d.get("total_plata"),
            err,
        ]

        for col, val in enumerate(row_values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = cell_font
            cell.border = border
            if fill:
                cell.fill = fill
            if col == money_col and isinstance(val, (int, float)):
                cell.number_format = money

    total_row = len(records) + 2
    ws.cell(row=total_row, column=7, value="TOTAL").font = bold_font
    total_cell = ws.cell(row=total_row, column=money_col,
                         value=f"=SUM(H2:H{total_row-1})")
    total_cell.font = bold_font
    total_cell.number_format = money

    wb.save(output_path)
    print(f"\n✅ Saved → {output_path}")
    print(f"   {len(records)} facturi")


def iter_inputs(input_path: Path):
    """Yield (label, extractor_fn) for each receipt to process.

    For a folder: yields (filename, lambda that calls extract_receipt).
    For a PDF:    yields (pdf_name — page N, lambda that calls extract_receipt_from_pil).
    """
    if input_path.is_dir():
        images = [f for f in sorted(input_path.iterdir())
                  if f.suffix.lower() in SUPPORTED_EXTENSIONS]
        if not images:
            print(f"No supported images found in {input_path}")
            return
        print(f"Found {len(images)} receipt image(s). Processing...\n")
        for img_path in images:
            yield img_path.name, img_path, None
    elif input_path.suffix.lower() == ".pdf":
        try:
            from pdf2image import convert_from_path
        except ImportError:
            print("pdf2image is not installed.")
            print("Install with:  pip install pdf2image")
            print("Also requires poppler — on macOS: brew install poppler")
            return
        pages = convert_from_path(str(input_path), dpi=200)
        print(f"Found {len(pages)} page(s) in '{input_path.name}'. Processing...\n")
        for n, pil_img in enumerate(pages, 1):
            yield f"{input_path.name} — page {n}", None, pil_img
    else:
        print(f"Unsupported input: {input_path}  (pass a folder or a .pdf file)")


def main():
    parser = argparse.ArgumentParser(description="Romanian fiscal document extractor via Claude Vision")
    parser.add_argument("--input", required=True, help="Folder with images, or a PDF file")
    parser.add_argument("--output", default="output files/output.xlsx", help="Output Excel filename")
    parser.add_argument("--mode", choices=["receipts", "facturi"], default="receipts",
                        help="Document type to process (default: receipts)")
    parser.add_argument("--skip-errors", action="store_true", help="Continue on parse failures")
    args = parser.parse_args()

    output = args.output
    if not os.path.dirname(output):
        output = os.path.join("output files", output)

    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise EnvironmentError("Set ANTHROPIC_API_KEY environment variable first.")
    if not os.environ.get("GOOGLE_VISION_API_KEY"):
        raise EnvironmentError("Set GOOGLE_VISION_API_KEY environment variable first.")

    client = anthropic.Anthropic(api_key=api_key)

    is_facturi = args.mode == "facturi"

    input_path = Path(args.input)
    inputs = list(iter_inputs(input_path))

    if not inputs:
        return

    records = []
    total = len(inputs)

    for i, (label, img_path, pil_img) in enumerate(inputs, 1):
        print(f"[{i}/{total}] {label} ... ", end="", flush=True)
        try:
            if is_facturi:
                data = extract_factura_from_pil(client, pil_img) if pil_img is not None else extract_factura(client, img_path)
                if not data.get("relevant", True):
                    vendor = (data.get("vendor") or {}).get("name", "?")
                    print(f"–  skipped ({vendor})")
                    continue
                records.append({"filename": label, "data": data, "error": ""})
                vendor = (data.get("vendor") or {}).get("name", "?")
                total_amount = data.get("total_plata", "?")
                print(f"✓  {vendor}  {total_amount} RON")
            else:
                data = extract_receipt_from_pil(client, pil_img) if pil_img is not None else extract_receipt(client, img_path)
                data["vat_breakdown"] = normalize_vat_by_rate(data.get("vat_breakdown", {}))
                warnings = validate_receipt(data)
                for w in warnings:
                    print(f"\n  ⚠  {w}", end="")
                records.append({"filename": label, "data": data, "error": "; ".join(warnings)})
                vendor = (data.get("vendor") or {}).get("name", "?")
                total_amount = data.get("total_with_vat", "?")
                print(f"✓  {vendor}  {total_amount} RON")
        except Exception as e:
            print(f"✗  ERROR: {e}")
            records.append({"filename": label, "data": {}, "error": str(e)})
            if not args.skip_errors:
                print("Stopping. Use --skip-errors to continue past failures.")
                break

    if records:
        if is_facturi:
            build_excel_facturi(records, output)
        else:
            build_excel(records, output)


if __name__ == "__main__":
    main()